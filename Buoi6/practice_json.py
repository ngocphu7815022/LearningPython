import openpyxl
import json

# Bai 1: Đọc file JSON
"""
with open("../data/info.json", "r") as f:
    data = json.load(f)

for key, value in data.items():
    print(f"{key}: {value}")


# Bai 2: Đọc dict sau đó ghi ra JSON
my_dict = {"name": "John", "age": 30, "city": "New York"}

with open("../data/person.json", "w") as file:
    json.dump(my_dict, file)


# Bai 3: Chỉnh sửa file JSON sau đó lưu vào 1 file JSON mới
with open("../data/students.json", "r") as f:
    data = json.load(f)


new_dict = {"name": "C", "score": 8}
data.append(new_dict)

with open("../data/students_updated.json", "w") as file:
    json.dump(data, file, indent=4, ensure_ascii=False)


# Bai 4: Đọc dữ liệu từ file Excel sau đó ghi ra JSON
wb = openpyxl.load_workbook("../data/employees.xlsx")
ws = wb.active

new_list = []
headers = ["name", "age", "city"]

for row in ws.iter_rows(min_row=2, values_only=True):
    new_list.append(dict(zip(headers, row)))

with open("../data/employees.json", "w") as file:
    json.dump(new_list, file, indent=4, ensure_ascii=False)
"""

# Bai 5: Đọc dữ liệu từ file JSON, sau đó in vào file excel
# Đọc file JSON
with open("../data/products.json", "r") as f:
    data = json.load(f)

# Ghi vào file Excel
wb = openpyxl.Workbook()
ws = wb.active

headers = ["product", "price"]
ws.append(headers)

for item in data:
    row = [item["product"], item["price"]]
    ws.append(row)

wb.save("../data/Bai12_Excel.xlsx")
# Đọc từ file Excel, lưu thành 1 file JSON mới
wb = openpyxl.load_workbook("../data/Bai12_Excel.xlsx")
ws = wb.active

new_list = []
headers = ["product", "price"]
for row in ws.iter_rows(min_row=2, values_only=True):
    new_list.append(dict(zip(headers, row)))

with open("../data/Bai12.json", "w") as file:
    json.dump(new_list, file, indent=4, ensure_ascii=False)
