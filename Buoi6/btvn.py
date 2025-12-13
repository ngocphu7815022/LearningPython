import openpyxl
import json

# Mở file
workbook = openpyxl.load_workbook("../data/Book1.xlsx")

# Chọn sheet theo tên
sheet = workbook["Sheet1"]

# Duyệt từng dòng và từng ô
# for row in sheet.iter_rows(values_only=True):
# print(row)
new_list = []
headers = ["A", "B", "C", "D"]

for row in sheet.iter_rows(min_row=2, values_only=True):
    new_list.append(dict(zip(headers, row)))

with open("report.json", "w", encoding="utf-8") as file:
    json.dump(new_list, file, ensure_ascii=False, indent=4)
