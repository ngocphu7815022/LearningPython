import openpyxl
import json


workbook = openpyxl.load_workbook("../data/employees.xlsx")

# File chỉ có 1 sheet
sheet = workbook.active


# Bai 1: In từng dòng dưới dạng tuple
for row in sheet.iter_rows(values_only=True):
    print(row)

"""
# Bai 2: In ra kết quả dạng list các tuple cho vùng B2:C4
data = []
for row in sheet["B2":"C4"]:
    data.append(tuple(cell.value for cell in row))

print(data)


# Bai 3: Chuyển dữ liệu từ Excel thành list of dict (lấy từ bài 1)
new_list = []
headers = ["name", "age", "city"]

for row in sheet.iter_rows(min_row=2, values_only=True):
    new_list.append(dict(zip(headers, row)))

# print(new_list)

# Bai 4: Lọc những người có age >= 22 (lấy dữ liệu từ bài 3)

list_bai4 = []
for item in new_list:
    if item["age"] >= 22:
        list_bai4.append(item)

# print(list_bai4)


# Bai 5: Ghi dữ liệu đã đọc ở bài 4 vào file excel mới

new_workbook = openpyxl.Workbook()

new_sheet = new_workbook.active

new_sheet.title = "Filtered Data"

# 2. Ghi header
headers1 = ["name", "age", "city"]
new_sheet.append(headers1)

for item in list_bai4:
    row = [item["name"], item["age"], item["city"]]
    new_sheet.append(row)


new_workbook.save("../data/filtered.xlsx")

# Bai 6: Ghi 1 dòng mới vào file bài 5 và tạo thành 1 file mới tên là update=
# Mở file
workbook = openpyxl.load_workbook("../data/filtered.xlsx")

# Chọn sheet theo tên
sheet = workbook.active

new_row = ["D", 30, "Can Tho"]

sheet.append(new_row)

workbook.save("../data/employees_updated.xlsx")
"""
# Bai 7: Đoc dữ liệu từ file JSON, sau đó in ra file Excel

# Đọc file JSON
with open("../data/products.json", "r") as file:
    data_from_json = json.load(file)


wb = openpyxl.Workbook()
ws = wb.active

headers = ["product", "price"]
ws.append(headers)

for item in data_from_json:
    row = [item["product"], item["price"]]
    ws.append(row)


wb.save("../data/products.xlsx")
